#!/usr/bin/env python
"""
Script to import inventory data from Excel file to database
"""
import os
import sys
import django
from django.db import transaction
from datetime import datetime

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'property_management.settings')
django.setup()

from openpyxl import load_workbook
from properties.models import HousingUnit, PropertyInventory # Assuming this is the correct path

def extract_date(date_str):
    """Convert date string to Python date object"""
    try:
        # Try multiple date formats
        for fmt in ['%B %d, %Y', '%b %d, %Y', '%m/%d/%Y', '%d/%m/%Y']:
            try:
                return datetime.strptime(date_str.strip(), fmt).date()
            except ValueError:
                continue
    except (AttributeError, ValueError):
        print(f"Warning: Could not parse date: '{date_str}'")
    return None

def import_excel_inventory(file_path):
    """Import inventory data from Excel file"""
    
    print(f"\n{'='*100}")
    print(f"IMPORTING INVENTORY DATA FROM: {file_path}")
    print(f"{'='*100}\n")
    
    # Open workbook
    try:
        wb = load_workbook(filename=file_path, data_only=True)
        sheet = wb.active
    except FileNotFoundError:
        print(f"ERROR: File not found at {file_path}")
        return
    
    # Extract header information (Rows 0-5)
    print("EXTRACTING HEADER INFORMATION...")
    
    # Using .cell(row, column).value syntax from openpyxl
    # Row 4: Occupant info
    occupant_name = sheet.cell(row=5, column=13).value  # M5 (openpyxl is 1-based)
    department = sheet.cell(row=5, column=34).value      # AH5
    section = sheet.cell(row=5, column=43).value         # AQ5
    job_title = sheet.cell(row=5, column=55).value       # BC5
    
    # Row 5: Housing unit info
    housing_unit_name = sheet.cell(row=6, column=6).value   # F6
    building = sheet.cell(row=6, column=15).value           # O6
    floor = sheet.cell(row=6, column=18).value              # R6
    unit_number = sheet.cell(row=6, column=23).value        # W6
    address = sheet.cell(row=6, column=32).value            # AF6
    
    # Row 1: Report date
    # Note: openpyxl cell coordinates are (row, column) and are 1-based.
    # xlrd is (row_index, col_index) and is 0-based.
    # The original `sheet.cell_value(1, 47)` corresponds to row 2, column 48 (AV2)
    date_str = sheet.cell(row=2, column=48).value  # "Petsa ng Pag-uulat: July 04, 2024"
    date_reported = None
    if date_str:
        # Extract date from string like "Petsa ng Pag-uulat: July 04, 2024"
        if ': ' in str(date_str):
            date_str = str(date_str).split(': ')[1]
        date_reported = extract_date(date_str)
    
    print(f"Occupant: {occupant_name}")
    print(f"Department: {department}")
    print(f"Section: {section}")
    print(f"Job Title: {job_title}")
    print(f"Housing Unit: {housing_unit_name}")
    print(f"Building: {building}")
    print(f"Address: {address}")
    print(f"Date Reported: {date_reported}\n")
    
    # Create or get HousingUnit
    print("CREATING/UPDATING HOUSING UNIT...")
    housing_unit, created = HousingUnit.objects.update_or_create(
        housing_unit_name=housing_unit_name,
        building=building,
        defaults={
            'occupant_name': occupant_name,
            'department': department,
            'section': section,
            'job_title': job_title,
            'date_reported': date_reported or datetime.now().date(),
            'floor': floor,
            'unit_number': unit_number,
            'address': address,
        }
    )
    
    if created:
        print(f"✓ Created new Housing Unit: {housing_unit}")
    else:
        print(f"✓ Updated existing Housing Unit: {housing_unit}")
    
    # Extract inventory items (Rows 9 onwards)
    print(f"\nEXTRACTING INVENTORY ITEMS...")
    print("-" * 100)
    
    inventory_count = 0
    skipped_count = 0
    
    # Column mapping
    col_date_acquired = 3
    col_qty = 7
    col_item_name = 9
    col_make = 32
    col_color = 37
    col_size = 42
    col_remarks = 52
    
    for row_idx in range(9, sheet.nrows):
        try:
            # Get basic fields
            date_acquired_val = sheet.cell_value(row_idx, col_date_acquired)
            qty = sheet.cell_value(row_idx, col_qty)
            item_name = sheet.cell_value(row_idx, col_item_name)
            
            # Skip empty rows
            if not item_name or item_name.strip() == '':
                continue
            
            # Convert date to year
            if isinstance(date_acquired_val, float):
                date_acquired = int(date_acquired_val)
            else:
                date_acquired = 2024  # Default year
            
            # Convert quantity to int
            if isinstance(qty, float):
                quantity = int(qty)
            else:
                quantity = 1
            
            # Get description fields
            make = sheet.cell_value(row_idx, col_make)
            color = sheet.cell_value(row_idx, col_color)
            size = sheet.cell_value(row_idx, col_size)
            remarks = sheet.cell_value(row_idx, col_remarks)
            
            # Create inventory item
            inventory_item = PropertyInventory.objects.create(
                housing_unit=housing_unit,
                date_acquired=date_acquired,
                quantity=quantity,
                item_name=item_name,
                make=make,
                color=color,
                size=size,
                remarks=remarks,
            )
            
            print(f"✓ Row {row_idx}: {item_name} (Qty: {quantity})")
            inventory_count += 1
            
        except Exception as e:
            print(f"✗ Row {row_idx}: Error - {str(e)}")
            skipped_count += 1
            continue
    
    print(f"\n{'='*100}")
    print(f"IMPORT COMPLETE")
    print(f"{'='*100}")
    print(f"Housing Units: 1")
    print(f"Inventory Items Created: {inventory_count}")
    print(f"Rows Skipped: {skipped_count}")
    print(f"Total Items: {inventory_count}")
    print(f"{'='*100}\n")

if __name__ == '__main__':
    # Find the Excel file
    excel_files = [f for f in os.listdir('.') if f.endswith('.xls') or f.endswith('.xlsx')]
    
    if not excel_files:
        print("ERROR: No Excel files found in current directory")
        sys.exit(1)
    
    file_path = excel_files[0]
    print(f"Found file: {file_path}")
    
    import_excel_inventory(file_path)
