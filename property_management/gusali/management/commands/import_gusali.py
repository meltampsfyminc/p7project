"""
Import building data from GUSALI REPORT.xlsx format

Usage:
    python manage.py import_gusali "path/to/GUSALI REPORT.xlsx"
    python manage.py import_gusali "path/to/GUSALI REPORT.xlsx" --force
    python manage.py import_gusali "path/to/GUSALI REPORT.xlsx" --clear
"""

from django.core.management.base import BaseCommand
from datetime import datetime, date
import openpyxl
import hashlib
import os
from decimal import Decimal
from gusali.models import Building, BuildingYearlyRecord
from properties.models import ImportedFile, Local


class Command(BaseCommand):
    help = 'Import building/gusali data from Excel file (GUSALI REPORT.xlsx format)'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to Excel file (.xlsx)')
        parser.add_argument('--force', action='store_true', help='Force import even if already imported')
        parser.add_argument('--clear', action='store_true', help='Clear existing buildings before import')
        parser.add_argument('--local', type=str, help='Local code to associate buildings with')

    def handle(self, *args, **options):
        file_path = options['file_path']
        force_import = options.get('force', False)
        local_code = options.get('local')
        
        # Check if file exists
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'File not found: {file_path}'))
            return
        
        # Open workbook
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error opening file: {str(e)}'))
            return
        
        # Calculate file hash for duplicate detection
        file_hash = self._calculate_file_hash(file_path)
        file_size = os.path.getsize(file_path)
        filename = os.path.basename(file_path)
        
        # Check if already imported (unless --force flag)
        if not force_import:
            existing_import = ImportedFile.objects.filter(file_hash=file_hash).first()
            if existing_import:
                self.stdout.write(self.style.WARNING(f'\n{"="*80}'))
                self.stdout.write(self.style.WARNING(f'FILE ALREADY IMPORTED'))
                self.stdout.write(self.style.WARNING(f'{"="*80}'))
                self.stdout.write(self.style.WARNING(f'Filename: {filename}'))
                self.stdout.write(self.style.WARNING(f'Previously imported on: {existing_import.imported_at}'))
                self.stdout.write(self.style.WARNING(f'Records imported: {existing_import.records_imported}'))
                self.stdout.write(self.style.WARNING(f'\nUse --force flag to re-import'))
                self.stdout.write(self.style.WARNING(f'{"="*80}\n'))
                return
        
        ws = wb.active
        
        self.stdout.write(self.style.SUCCESS(f'\n{"="*80}'))
        self.stdout.write(self.style.SUCCESS(f'IMPORTING GUSALI (BUILDING) DATA FROM: {filename}'))
        self.stdout.write(self.style.SUCCESS(f'{"="*80}\n'))
        
        # Get local if specified
        local_obj = None
        if local_code:
            local_obj = Local.objects.filter(lcode=local_code).first()
            if local_obj:
                self.stdout.write(self.style.SUCCESS(f'Associated with Local: {local_obj}'))
            else:
                self.stdout.write(self.style.WARNING(f'Local code not found: {local_code}'))
        
        # Extract year from header (Row 4, Column 7)
        year_covered = self._safe_int(ws.cell(row=4, column=7).value, 2024)
        self.stdout.write(f'Year Covered: {year_covered}')
        
        # Clear existing if requested
        if options['clear']:
            self.stdout.write(self.style.WARNING('Clearing existing buildings...'))
            Building.objects.all().delete()
        
        # Extract building data (starting from row 7)
        self.stdout.write(self.style.SUCCESS(f'\nEXTRACTING BUILDING DATA...'))
        self.stdout.write('-' * 80)
        
        building_count = 0
        record_count = 0
        skipped_count = 0
        
        # Column mapping based on GUSALI REPORT.xlsx analysis
        # Row 6 contains headers: CODE, BLDG, ORIGINAL COST, CLASS, etc.
        
        for row_num in range(7, ws.max_row + 1):
            try:
                # Read all cells in row
                code = ws.cell(row=row_num, column=1).value
                name = ws.cell(row=row_num, column=2).value
                
                # Skip rows without building code or name
                if not code or not name:
                    continue
                
                # Skip if code is not a valid building code
                if code not in ['A', 'B', 'C', 'D']:
                    continue
                
                original_cost = self._safe_decimal(ws.cell(row=row_num, column=3).value)
                classification = str(ws.cell(row=row_num, column=4).value or '')
                donation_status = str(ws.cell(row=row_num, column=5).value or '')
                donation_date = self._parse_date(ws.cell(row=row_num, column=6).value)
                ownership_date = self._parse_date(ws.cell(row=row_num, column=7).value)
                constructor = str(ws.cell(row=row_num, column=8).value or '')
                capacity = self._safe_int(ws.cell(row=row_num, column=9).value, None)
                cost_last_year = self._safe_decimal(ws.cell(row=row_num, column=10).value)
                
                # Additions
                construction_cost = self._safe_decimal(ws.cell(row=row_num, column=11).value)
                renovation_cost = self._safe_decimal(ws.cell(row=row_num, column=12).value)
                general_repair_cost = self._safe_decimal(ws.cell(row=row_num, column=13).value)
                other_additions = self._safe_decimal(ws.cell(row=row_num, column=14).value)
                total_added = self._safe_decimal(ws.cell(row=row_num, column=15).value)
                
                remarks = str(ws.cell(row=row_num, column=16).value or '')
                
                # Broken/removed
                broken_part = str(ws.cell(row=row_num, column=17).value or '')
                broken_cost = self._safe_decimal(ws.cell(row=row_num, column=18).value)
                
                # Total
                year_end_total = self._safe_decimal(ws.cell(row=row_num, column=19).value)
                
                # Determine if donated
                is_donated = 'HANDOG' in donation_status.upper() if donation_status else False
                
                # Create or update Building
                building, created = Building.objects.update_or_create(
                    name=name,
                    local=local_obj,
                    defaults={
                        'code': code,
                        'classification': classification,
                        'is_donated': is_donated,
                        'donation_date': donation_date,
                        'ownership_date': ownership_date,
                        'constructor': constructor,
                        'capacity': capacity,
                        'original_cost': original_cost,
                        'current_total_cost': year_end_total or cost_last_year or original_cost,
                        'year_covered': year_covered,
                        'remarks': remarks,
                    }
                )
                
                action = 'Created' if created else 'Updated'
                self.stdout.write(f'✓ {action}: {code} - {name} (₱{year_end_total or cost_last_year:,.2f})')
                building_count += 1
                
                # Create yearly record
                yearly_record, yr_created = BuildingYearlyRecord.objects.update_or_create(
                    building=building,
                    year=year_covered,
                    defaults={
                        'cost_last_year': cost_last_year,
                        'construction_cost': construction_cost,
                        'renovation_cost': renovation_cost,
                        'general_repair_cost': general_repair_cost,
                        'other_additions_cost': other_additions,
                        'broken_removed_part': broken_part,
                        'broken_removed_cost': broken_cost,
                        'remarks': remarks,
                    }
                )
                record_count += 1
                
            except Exception as e:
                self.stdout.write(self.style.ERROR(f'✗ Row {row_num}: Error - {str(e)}'))
                skipped_count += 1
                continue
        
        # Summary
        self.stdout.write(self.style.SUCCESS(f'\n{"="*80}'))
        self.stdout.write(self.style.SUCCESS('IMPORT COMPLETE'))
        self.stdout.write(self.style.SUCCESS(f'{"="*80}'))
        self.stdout.write(self.style.SUCCESS(f'Buildings Created/Updated: {building_count}'))
        self.stdout.write(self.style.SUCCESS(f'Yearly Records Created: {record_count}'))
        self.stdout.write(self.style.WARNING(f'Rows Skipped: {skipped_count}'))
        self.stdout.write(self.style.SUCCESS(f'{"="*80}\n'))
        
        # Save import record
        status = 'partial' if skipped_count > 0 else 'success'
        ImportedFile.objects.update_or_create(
            file_hash=file_hash,
            defaults={
                'filename': filename,
                'file_size': file_size,
                'records_imported': building_count,
                'status': status,
                'error_message': f'Skipped {skipped_count} rows' if skipped_count > 0 else ''
            }
        )
        
        self.stdout.write(self.style.SUCCESS(f'Import record saved. File Hash: {file_hash[:16]}...'))

    def _calculate_file_hash(self, file_path):
        """Calculate SHA256 hash of file for duplicate detection"""
        sha256_hash = hashlib.sha256()
        with open(file_path, "rb") as f:
            for byte_block in iter(lambda: f.read(4096), b""):
                sha256_hash.update(byte_block)
        return sha256_hash.hexdigest()

    def _safe_decimal(self, value, default=Decimal('0')):
        """Safely convert value to Decimal"""
        if value is None:
            return default
        try:
            # Handle strings like "WALA PO" or other non-numeric
            if isinstance(value, str):
                value = value.strip()
                if not value or value.upper() in ['WALA', 'WALA PO', '-', 'N/A', '']:
                    return default
                value = value.replace(',', '')
            return Decimal(str(value))
        except:
            return default

    def _safe_int(self, value, default=None):
        """Safely convert value to int"""
        if value is None:
            return default
        try:
            return int(float(value))
        except:
            return default

    def _parse_date(self, value):
        """Parse date from various formats"""
        if value is None:
            return None
        
        # If already a datetime
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        
        # If string
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            
            # Try various formats
            formats = [
                '%B %d, %Y',      # April 29, 1989
                '%B %Y',          # April 1989
                '%m/%d/%Y',       # 04/29/1989
                '%d/%m/%Y',       # 29/04/1989
                '%Y-%m-%d',       # 1989-04-29
            ]
            
            for fmt in formats:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        
        return None
